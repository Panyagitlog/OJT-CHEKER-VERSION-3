"""
core/report_generator.py - Excel Report Generator
OJT Checker System

Collects ValidationReport objects and writes a formatted
Excel summary report using openpyxl with color-coded cells.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# Color constants (openpyxl ARGB hex)
# ─────────────────────────────────────────────
COLOR_PASS_FILL    = "FF92D050"   # Light green
COLOR_FAIL_FILL    = "FFFF0000"   # Red
COLOR_WARN_FILL    = "FFFFC000"   # Amber
COLOR_HEADER_FILL  = "FF1F4E79"   # Dark blue
COLOR_ALT_ROW      = "FFD6E4F0"   # Light blue tint
COLOR_WHITE        = "FFFFFFFF"
COLOR_BLACK        = "FF000000"
COLOR_HEADER_FONT  = "FFFFFFFF"   # White


class ReportGenerator:
    """
    Builds a formatted Excel report from processed ValidationReport objects.

    Usage:
        gen = ReportGenerator()
        gen.add(validation_report)       # call for each PDF
        path = gen.save(output_dir)
        print(f"Report saved: {path}")
    """

    def __init__(self):
        self._rows: list[dict] = []
        self._stats = {
            "total": 0,
            "pass": 0,
            "fail": 0,
            "error": 0,
            "corrections": 0,
        }

    # ─────────────────────────────────────────
    # Public API
    # ─────────────────────────────────────────

    def add(self, report) -> None:
        """Add one ValidationReport to the report."""
        row = self._report_to_row(report)
        self._rows.append(row)
        self._update_stats(report)

    def save(self, output_dir: Path, filename: Optional[str] = None) -> Path:
        """
        Save the Excel report to output_dir.
        Returns the path to the saved file.
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        if filename is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = config.REPORT_FILENAME_TEMPLATE.format(timestamp=ts)

        output_path = output_dir / filename

        wb = openpyxl.Workbook()
        ws_detail = wb.active
        ws_detail.title = "Validation Results"

        self._write_detail_sheet(ws_detail)

        ws_summary = wb.create_sheet("Summary")
        self._write_summary_sheet(ws_summary)

        try:
            wb.save(str(output_path))
            logger.info(f"Report saved: {output_path}")
            return output_path
        except Exception as exc:
            logger.error(f"Failed to save report: {exc}")
            raise

    @property
    def stats(self) -> dict:
        return self._stats.copy()

    # ─────────────────────────────────────────
    # Sheet builders
    # ─────────────────────────────────────────

    def _write_detail_sheet(self, ws) -> None:
        """Write the per-PDF validation results sheet."""
        # Header row
        headers = config.REPORT_COLUMNS
        ws.append(headers)
        self._style_header_row(ws, 1, len(headers))

        # Data rows
        for i, row in enumerate(self._rows, start=2):
            values = [row.get(col, "") for col in headers]
            ws.append(values)
            self._style_data_row(ws, i, row)

        # Auto-size columns
        self._autofit_columns(ws, headers)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _write_summary_sheet(self, ws) -> None:
        """Write the summary statistics sheet."""
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        title_font = Font(bold=True, size=14, color=COLOR_HEADER_FONT)
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER_FILL)

        ws["A1"] = "OJT Checker System – Processing Summary"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")

        ws["A2"] = "Generated"
        ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        rows = [
            ("Total PDFs Processed", self._stats["total"]),
            ("PASS", self._stats["pass"]),
            ("FAIL", self._stats["fail"]),
            ("Errors", self._stats["error"]),
            ("Total Corrections Made", self._stats["corrections"]),
            ("Pass Rate (%)", self._pass_rate()),
        ]

        for r_idx, (label, value) in enumerate(rows, start=4):
            ws.cell(r_idx, 1, label).font = Font(bold=True)
            cell = ws.cell(r_idx, 2, value)
            if label == "PASS":
                cell.fill = PatternFill("solid", fgColor=COLOR_PASS_FILL)
            elif label == "FAIL":
                cell.fill = PatternFill("solid", fgColor=COLOR_FAIL_FILL)

    # ─────────────────────────────────────────
    # Row conversion
    # ─────────────────────────────────────────

    def _report_to_row(self, report) -> dict:
        """Convert a ValidationReport to a flat dict for one Excel row."""
        fr = report.field_results

        def field_status(key: str) -> str:
            return fr[key].status if key in fr else "N/A"

        return {
            "PDF Name":          report.pdf_name,
            "Enrollment Number": report.enrollment_no,
            "Student Name":      fr.get("student_name", {}).ocr_value if "student_name" in fr else "",
            "Course Match":      field_status("course_name"),
            "Name Match":        field_status("student_name"),
            "OJT Period Match":  field_status("ojt_period"),
            "Stamp Status":      report.stamp_status,
            "Corrections Made":  report.corrections_made,
            "Final Status":      report.final_status,
            "Error Notes":       report.error or "",
        }

    def _update_stats(self, report) -> None:
        self._stats["total"] += 1
        status = report.final_status
        if status == "PASS":
            self._stats["pass"] += 1
        elif status == "ERROR":
            self._stats["error"] += 1
        else:
            self._stats["fail"] += 1
        self._stats["corrections"] += report.corrections_made

    def _pass_rate(self) -> str:
        total = self._stats["total"]
        if total == 0:
            return "0.00%"
        rate = (self._stats["pass"] / total) * 100
        return f"{rate:.2f}%"

    # ─────────────────────────────────────────
    # Styling helpers
    # ─────────────────────────────────────────

    def _style_header_row(self, ws, row_num: int, col_count: int) -> None:
        fill = PatternFill("solid", fgColor=COLOR_HEADER_FILL)
        font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color=COLOR_BLACK)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in range(1, col_count + 1):
            cell = ws.cell(row_num, col)
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = border

        ws.row_dimensions[row_num].height = 30

    def _style_data_row(self, ws, row_num: int, row_data: dict) -> None:
        status = row_data.get("Final Status", "")
        bg_color = COLOR_WHITE

        if status == "PASS":
            bg_color = COLOR_PASS_FILL
        elif status in ("FAIL", "ERROR"):
            bg_color = COLOR_FAIL_FILL
        elif row_num % 2 == 0:
            bg_color = COLOR_ALT_ROW

        fill = PatternFill("solid", fgColor=bg_color)
        align = Alignment(vertical="center", wrap_text=False)

        cols = len(config.REPORT_COLUMNS)
        for col in range(1, cols + 1):
            cell = ws.cell(row_num, col)
            cell.fill = fill
            cell.alignment = align

            # Color individual status columns
            col_name = config.REPORT_COLUMNS[col - 1]
            cell_val = str(cell.value or "")
            if cell_val == "PASS":
                cell.fill = PatternFill("solid", fgColor=COLOR_PASS_FILL)
            elif cell_val in ("FAIL", "MISSING STAMP", "MISSING"):
                cell.fill = PatternFill("solid", fgColor=COLOR_FAIL_FILL)
            elif cell_val == "MISSING STAMP":
                cell.fill = PatternFill("solid", fgColor=COLOR_WARN_FILL)

    def _autofit_columns(self, ws, headers: list[str]) -> None:
        for i, header in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            max_len = len(str(header))
            for row in ws.iter_rows(
                min_col=i, max_col=i, min_row=2, max_row=ws.max_row
            ):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            # Cap column width
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
